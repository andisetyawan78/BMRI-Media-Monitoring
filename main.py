"""
Bank Mandiri Media Monitoring
==============================
Otomatis setiap hari jam 07:00 WIB via GitHub Actions.
Alur: Gmail (Talkwalker alerts) → Parse → Analisis AI → Grafik → Telegram + Email

Diperlukan environment variables (simpan sebagai GitHub Secrets):
  GMAIL_CREDENTIALS_JSON   : isi file credentials.json Google OAuth2 (base64-encoded)
  GMAIL_TOKEN_JSON         : isi file token.json Google OAuth2 (base64-encoded)
  ANTHROPIC_API_KEY        : API key Anthropic Claude
  TELEGRAM_BOT_TOKEN       : token bot Telegram
  TELEGRAM_CHAT_ID         : chat_id tujuan Telegram
  EMAIL_SENDER             : alamat Gmail pengirim (misal: sivbmri@gmail.com)
  EMAIL_APP_PASSWORD       : Gmail App Password (bukan password biasa)
  EMAIL_RECIPIENT          : sivbmri@gmail.com
"""

import os
import base64
import json
import re
import smtplib
import tempfile
import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
from html.parser import HTMLParser

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import anthropic

# ─────────────────────────────────────────────
# KONFIGURASI
# ─────────────────────────────────────────────
TALKWALKER_SENDER    = "alerts@talkwalker.com"
GOOGLE_ALERTS_SENDER = "googlealerts-noreply@google.com"
MONITOR_KEYWORDS     = ["Bank Mandiri", "BMRI"]
EMAIL_RECIPIENT   = os.environ.get("EMAIL_RECIPIENT", "sivbmri@gmail.com")
EMAIL_SENDER      = os.environ.get("EMAIL_SENDER", "sivbmri@gmail.com")
EMAIL_APP_PASS    = os.environ.get("EMAIL_APP_PASSWORD", "")
TG_TOKEN          = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TG_CHAT_ID        = os.environ.get("TELEGRAM_CHAT_ID", "")
ANTHROPIC_KEY     = os.environ.get("ANTHROPIC_API_KEY", "")
LISTENNOTES_KEY   = os.environ.get("LISTENNOTES_API_KEY", "")

# Domain e-commerce yang dideteksi otomatis dari Google Alerts
E_COMMERCE_DOMAINS = {
    "tokopedia.com", "shopee.co.id", "shopee.com", "lazada.co.id",
    "bukalapak.com", "blibli.com", "jd.id", "traveloka.com",
    "tiket.com", "zalora.co.id", "bhinneka.com",
}
GMAIL_SCOPES      = ["https://www.googleapis.com/auth/gmail.readonly"]


# ─────────────────────────────────────────────
# 1. AUTENTIKASI GMAIL
# ─────────────────────────────────────────────
def get_gmail_service():
    """Buat Gmail API service dari credentials yang tersimpan di env vars."""
    creds = None

    # Decode credentials dari env (base64)
    creds_b64   = os.environ.get("GMAIL_CREDENTIALS_JSON", "")
    token_b64   = os.environ.get("GMAIL_TOKEN_JSON", "")

    def safe_b64decode(s):
        """Decode base64 dengan auto-fix padding dan strip whitespace."""
        s = s.strip().replace("\n", "").replace("\r", "").replace(" ", "")
        # Tambah padding jika kurang
        missing = len(s) % 4
        if missing:
            s += "=" * (4 - missing)
        return base64.b64decode(s)

    with tempfile.TemporaryDirectory() as tmpdir:
        creds_path  = os.path.join(tmpdir, "credentials.json")
        token_path  = os.path.join(tmpdir, "token.json")

        if creds_b64:
            with open(creds_path, "w") as f:
                f.write(safe_b64decode(creds_b64).decode())

        if token_b64:
            with open(token_path, "w") as f:
                f.write(safe_b64decode(token_b64).decode())
            creds = Credentials.from_authorized_user_file(token_path, GMAIL_SCOPES)

        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())

        service = build("gmail", "v1", credentials=creds)
    return service


# ─────────────────────────────────────────────
# 2. AMBIL EMAIL TALKWALKER KEMARIN
# ─────────────────────────────────────────────
def get_yesterday_range():
    """Return (after_str, before_str) dalam format YYYY/MM/DD untuk kemarin."""
    today     = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)
    return yesterday.strftime("%Y/%m/%d"), today.strftime("%Y/%m/%d")


def fetch_talkwalker_emails(service):
    """Ambil semua email Talkwalker dari kemarin, return list HTML bodies."""
    after, before = get_yesterday_range()
    query = f"from:{TALKWALKER_SENDER} after:{after} before:{before}"
    print(f"[Gmail] Query: {query}")

    result   = service.users().messages().list(userId="me", q=query, maxResults=50).execute()
    messages = result.get("messages", [])
    print(f"[Gmail] Ditemukan {len(messages)} pesan")

    html_bodies = []
    for msg_meta in messages:
        msg = service.users().messages().get(
            userId="me", id=msg_meta["id"], format="full"
        ).execute()
        html = _extract_html(msg)
        if html:
            html_bodies.append(html)
    return html_bodies


def fetch_google_alerts_emails(service):
    """Ambil semua email Google Alerts dari kemarin, return list HTML bodies."""
    after, before = get_yesterday_range()
    query = f"from:{GOOGLE_ALERTS_SENDER} after:{after} before:{before}"
    print(f"[Google Alerts] Query: {query}")

    result   = service.users().messages().list(userId="me", q=query, maxResults=50).execute()
    messages = result.get("messages", [])
    print(f"[Google Alerts] Ditemukan {len(messages)} pesan")

    html_bodies = []
    for msg_meta in messages:
        msg = service.users().messages().get(
            userId="me", id=msg_meta["id"], format="full"
        ).execute()
        html = _extract_html(msg)
        if html:
            html_bodies.append(html)
    return html_bodies


def _extract_html(msg):
    """Ekstrak HTML body dari pesan Gmail."""
    payload = msg.get("payload", {})
    return _walk_parts(payload)


def _walk_parts(part):
    mime = part.get("mimeType", "")
    body = part.get("body", {})
    data = body.get("data", "")

    if mime == "text/html" and data:
        return base64.urlsafe_b64decode(data).decode("utf-8", errors="replace")

    for sub in part.get("parts", []):
        result = _walk_parts(sub)
        if result:
            return result
    return None


# ─────────────────────────────────────────────
# 3. PARSE HTML → DAFTAR ARTIKEL
# ─────────────────────────────────────────────
class _HTMLStripper(HTMLParser):
    def __init__(self):
        super().__init__()
        self.fed = []
        self._skip = False

    def handle_starttag(self, tag, attrs):
        if tag in ("script", "style"):
            self._skip = True

    def handle_endtag(self, tag):
        if tag in ("script", "style"):
            self._skip = False

    def handle_data(self, d):
        if not self._skip:
            self.fed.append(d)

    def get_data(self):
        return "\n".join(self.fed)


def html_to_text(html):
    s = _HTMLStripper()
    s.feed(html)
    text = s.get_data()
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r" {2,}", " ", text)
    return text.strip()


def extract_talkwalker_links(html):
    """
    Ekstrak semua link artikel dari HTML Talkwalker email.
    Returns: list of (anchor_normalized, anchor_words_set, url)

    Talkwalker membungkus link dalam redirect URL-nya sendiri dengan
    berbagai format parameter (url=, u=, link=, dst.), sehingga kita
    coba semua kemungkinan untuk mengekstrak URL artikel asli.
    """
    from urllib.parse import unquote
    tag_re  = re.compile(r"<[^>]+>")
    link_re = re.compile(
        r'<a[^>]+href=["\']([^"\']+)["\'][^>]*>(.*?)</a>',
        re.DOTALL | re.IGNORECASE,
    )
    SKIP_INTERNAL = [
        "unsubscribe", "mailto:", "javascript:", "tell-a-friend",
        "manage-alert", "/account", "delete-alert",
        "facebook.com", "linkedin.com",
    ]

    links = []  # list of (anchor_normalized, anchor_words_set, url)

    for m in link_re.finditer(html):
        url    = m.group(1).strip()
        anchor = tag_re.sub("", m.group(2)).strip()
        anchor = re.sub(r"\s+", " ", anchor)

        if not url.startswith("http") or not anchor or len(anchor) < 6:
            continue
        if any(s in url.lower() for s in SKIP_INTERNAL):
            continue

        # Talkwalker redirect → coba berbagai format parameter
        if "talkwalker.com" in url.lower():
            # Format 1: ?url=, &u=, &link=, &href=, &redirect=, &r=, &go=
            redirect_m = re.search(
                r"[?&](?:url|u|link|href|target|dest|redirect|r|go|to|click)"
                r"=([^&\s\"'<>]+)",
                url, re.IGNORECASE,
            )
            if redirect_m:
                candidate = unquote(redirect_m.group(1))
                if candidate.startswith("http"):
                    url = candidate
                else:
                    continue
            else:
                # Format 2: URL non-Talkwalker tersembunyi di path/query
                direct_m = re.search(
                    r"https?://(?!(?:www\.)?talkwalker)[^\s\"'&<>]+",
                    url,
                )
                url = direct_m.group(0) if direct_m else None
                if not url:
                    continue

        # Abaikan anchor yang terlalu pendek/generik (tombol "Read more", dll.)
        words = anchor.lower().split()
        if len(words) < 3:
            continue

        anchor_norm  = re.sub(r"[^\w\s]", "", anchor.lower()).strip()[:120]
        anchor_words = set(words)
        links.append((anchor_norm, anchor_words, url))

    return links


def find_article_link(title, source, links):
    """
    Cari URL terbaik untuk sebuah artikel dari daftar links Talkwalker.
    Strategi (berurutan):
      1. Exact match setelah normalisasi
      2. Prefix match (18 karakter pertama)
      3. Word-overlap >= 65% kata judul ada di anchor
      4. Fallback: domain sumber cocok dan unik
    """
    if not links:
        return ""

    title_key   = re.sub(r"[^\w\s]", "", title.lower()).strip()
    title_words = set(title.lower().split())

    # 1. Exact
    for anorm, _, url in links:
        if anorm == title_key:
            return url

    # 2. Prefix match (18 karakter)
    if len(title_key) >= 12:
        prefix = title_key[:18]
        for anorm, _, url in links:
            if anorm.startswith(prefix):
                return url

    # 3. Word-overlap
    best_url, best_overlap = "", 0
    for _, awords, url in links:
        if len(title_words) < 3:
            continue
        overlap = len(title_words & awords)
        ratio   = overlap / len(title_words)
        if ratio >= 0.60 and overlap > best_overlap:
            best_overlap = overlap
            best_url = url
    if best_url:
        return best_url

    # 4. Source domain fallback (hanya jika unik)
    if source:
        src_clean = re.sub(r"^www\.", "", source.lower()).split("/")[0]
        matches = [url for _, _, url in links
                   if src_clean.replace(".", "") in url.lower().replace(".", "")]
        if len(matches) == 1:
            return matches[0]

    return ""


def parse_articles(html_bodies):
    """
    Parse artikel dari HTML email Talkwalker.
    Strategi: cari baris tanggal/sumber (format Talkwalker), lalu scan mundur
    untuk menemukan judul dan snippet.
    Tidak menggunakan filter keyword karena Talkwalker sudah memfilter untuk Bank Mandiri.
    Return list of dict: {title, snippet, date, source, media_type}
    """
    articles = []

    # Format Talkwalker: "15/06/26 05:34 | Indonesia | katadata.co.id"
    # atau sumber di baris berikutnya: "15/06/26 05:34 | Indonesia |"
    date_pattern = re.compile(
        r"^(\d{2}/\d{2}/\d{2})\s+\d{2}:\d{2}\s*\|\s*([\w][\w\s]*)\|\s*(.*)$"
    )
    # Fallback: format tanpa jam "DD/MM/YY | Country | source"
    date_pattern_nohour = re.compile(
        r"^(\d{2}/\d{2}/\d{2,4})\s*\|\s*([\w][\w\s]*)\|\s*(.*)$"
    )
    # Baris noise yang harus diabaikan saat scan mundur
    NOISE = {
        "Tweet", "Show less", "Tell a Friend", "liking", "following",
        "News", "Blogs", "Twitter", "new results", "Delete Alert",
        "Manage Alerts", "Create Alert", "Unsubscribe", "View in browser",
    }
    NOISE_RE = re.compile(r"^\d+\s+new results?$", re.IGNORECASE)

    for html in html_bodies:
        # Ekstrak semua link dari HTML sebelum di-strip ke plain text
        tw_links = extract_talkwalker_links(html)

        text = html_to_text(html)
        # Collapse tab/spasi ganda tapi jaga newline
        text = re.sub(r"[ \t]{2,}", " ", text)

        all_lines = text.splitlines()
        lines = [l.strip() for l in all_lines]  # semua baris (termasuk kosong)

        # Deteksi media_type dari 40 baris pertama yang tidak kosong
        media_type = "News"
        non_empty_head = [l for l in lines[:60] if l][:40]
        for line in non_empty_head:
            ll = line.lower()
            if "blogs" in ll:
                media_type = "Blog"
                break
            elif "twitter" in ll:
                media_type = "Twitter"
                break

        # Indeks baris tidak-kosong untuk navigasi
        nonempty_lines = [(i, l) for i, l in enumerate(lines) if l]

        for idx, (line_idx, line) in enumerate(nonempty_lines):
            # Coba cocokkan pola tanggal
            m = date_pattern.match(line) or date_pattern_nohour.match(line)
            if not m:
                continue

            date = m.group(1)
            # group(3) = sumber jika ada di baris yang sama
            inline_source = m.group(3).strip() if m.lastindex >= 3 else ""

            # Jika sumber kosong, cek baris non-kosong berikutnya
            source = inline_source
            if not source and idx + 1 < len(nonempty_lines):
                next_line = nonempty_lines[idx + 1][1]
                # Validasi: terlihat seperti domain (ada titik, tidak terlalu panjang)
                if re.match(r"^[\w\-]+\.[\w\.\-]{2,}$", next_line) and len(next_line) < 60:
                    source = next_line

            if not source:
                continue

            # Scan mundur untuk menemukan judul dan snippet
            title   = ""
            snippet = ""
            candidates = []  # (baris non-kosong sebelum date line)

            for back_idx in range(idx - 1, max(idx - 20, -1), -1):
                prev_line = nonempty_lines[back_idx][1]

                # Hentikan jika ketemu baris tanggal lain (artikel sebelumnya)
                if date_pattern.match(prev_line) or date_pattern_nohour.match(prev_line):
                    break
                # Hentikan jika ketemu noise header
                if prev_line in NOISE or NOISE_RE.match(prev_line):
                    break
                # Lewati baris sangat pendek (1-3 karakter) — biasanya artefak HTML
                if len(prev_line) <= 3:
                    continue

                candidates.insert(0, prev_line)
                if len(candidates) >= 5:
                    break

            # Bersihkan candidates dari noise
            candidates = [
                c for c in candidates
                if c not in NOISE and not NOISE_RE.match(c)
            ]

            if not candidates:
                continue

            # Gabungkan kandidat pendek berturut-turut (kata terpotong akibat HTML tag)
            # Contoh: ["Bank", "Mandiri", "Gelar Program..."] → "Bank Mandiri Gelar Program..."
            title = candidates[0]
            rest  = candidates[1:]
            while len(title.split()) <= 2 and rest:
                title = title + " " + rest[0]
                rest  = rest[1:]
            snippet = " ".join(rest[:2])[:300]

            # Filter judul yang tidak masuk akal
            if len(title) < 8:
                continue
            skip_titles = {"tell a friend", "new results", "delete alert",
                           "manage alerts", "create alert", "unsubscribe"}
            if any(s in title.lower() for s in skip_titles):
                continue

            # Cari URL artikel menggunakan word-overlap matching
            article_link = find_article_link(title, source.strip(), tw_links)

            articles.append({
                "title"      : title,
                "snippet"    : snippet,
                "date"       : date,
                "source"     : source.strip(),
                "media_type" : media_type,
                "link"       : article_link,
            })

    # Deduplikasi berdasarkan judul
    seen   = set()
    unique = []
    for a in articles:
        key = a["title"][:60].lower()
        if key not in seen:
            seen.add(key)
            unique.append(a)

    n_with_link = sum(1 for a in unique if a.get("link"))
    print(f"[Parser] Total artikel unik: {len(unique)} | Dengan link: {n_with_link}")
    return unique


# ─────────────────────────────────────────────
# 3b. PARSE GOOGLE ALERTS
# ─────────────────────────────────────────────
def parse_google_alerts(html_bodies):
    """
    Parse artikel dari Google Alerts email HTML.
    Struktur: <h3><a href="google-redirect">Judul</a></h3>
              <div><a href="...">Nama Sumber</a> - Tanggal</div>
              <div>Snippet...</div>
    Return list of dict: {title, snippet, date, source, media_type}
    """
    from urllib.parse import unquote

    articles   = []
    tag_re     = re.compile(r"<[^>]+>")
    # Setiap blok artikel diawali <h3> dan diikuti konten sampai <h3> berikutnya
    block_re   = re.compile(
        r"<h3[^>]*>.*?<a[^>]+href=\"([^\"]+)\"[^>]*>(.*?)</a>.*?</h3>(.*?)(?=<h3|</table)",
        re.DOTALL | re.IGNORECASE,
    )
    # Link sumber di div setelah <h3>
    src_link_re = re.compile(
        r"<a[^>]+href=\"([^\"]+)\"[^>]*>(.*?)</a>",
        re.DOTALL | re.IGNORECASE,
    )
    # Ekstrak URL asli dari Google redirect (?url=...)
    gurl_re = re.compile(r"[?&]url=([^&]+)", re.IGNORECASE)

    for html in html_bodies:
        for m in block_re.finditer(html):
            raw_url   = m.group(1)
            title_raw = m.group(2)
            rest_html = m.group(3)

            title = tag_re.sub("", title_raw).strip()
            title = re.sub(r"\s+", " ", title)
            if not title or len(title) < 8:
                continue

            # Ekstrak domain dari Google redirect URL
            gurl_m = gurl_re.search(raw_url)
            actual_url = unquote(gurl_m.group(1)) if gurl_m else raw_url
            domain_m   = re.search(r"https?://(?:www\.)?([^/?#]+)", actual_url)
            domain     = domain_m.group(1) if domain_m else ""

            # Nama sumber dari link pertama di blok setelah <h3>
            src_m    = src_link_re.search(rest_html)
            src_name = tag_re.sub("", src_m.group(2)).strip() if src_m else domain

            # Snippet: teks bersih dari sisa blok
            snippet = tag_re.sub(" ", rest_html)
            snippet = re.sub(r"\s+", " ", snippet).strip()[:300]

            if not src_name:
                src_name = domain

            # Deteksi e-commerce dari domain
            is_ecommerce = any(d in domain for d in E_COMMERCE_DOMAINS)
            media_type   = "E-Commerce" if is_ecommerce else "Google Alerts"

            articles.append({
                "title"     : title,
                "snippet"   : snippet,
                "date"      : "",
                "source"    : src_name,
                "media_type": media_type,
                "link"      : actual_url,
            })

    # Deduplikasi berdasarkan judul
    seen   = set()
    unique = []
    for a in articles:
        key = a["title"][:60].lower()
        if key not in seen:
            seen.add(key)
            unique.append(a)

    print(f"[Google Alerts] Total artikel unik: {len(unique)}")
    return unique


# ─────────────────────────────────────────────
# 3c. FETCH PODCAST (LISTENNOTES API)
# ─────────────────────────────────────────────
def fetch_podcast_mentions():
    """
    Cari sebutan Bank Mandiri / BMRI di episode podcast via ListenNotes API.
    Memerlukan env LISTENNOTES_API_KEY (daftar gratis di listennotes.com/api).
    """
    if not LISTENNOTES_KEY:
        print("[Podcast] LISTENNOTES_API_KEY tidak diset, skip.")
        return []

    yesterday  = datetime.date.today() - datetime.timedelta(days=1)
    pub_after  = int(datetime.datetime.combine(
        yesterday, datetime.time.min, tzinfo=datetime.timezone.utc
    ).timestamp())

    articles = []
    for keyword in MONITOR_KEYWORDS:
        try:
            resp = requests.get(
                "https://listen-api.listennotes.com/api/v2/search",
                headers={"X-ListenAPI-Key": LISTENNOTES_KEY},
                params={
                    "q"              : keyword,
                    "type"           : "episode",
                    "published_after": pub_after,
                    "safe_mode"      : 0,
                    "only_in"        : "title,description",
                },
                timeout=20,
            )
            if resp.status_code == 401:
                print("[Podcast] API key tidak valid.")
                break
            if resp.status_code != 200:
                print(f"[Podcast] ListenNotes error: {resp.status_code}")
                continue

            for ep in resp.json().get("results", []):
                title        = ep.get("title_original", "").strip()
                podcast_name = ep.get("podcast", {}).get("title_original", "Podcast")
                snippet      = re.sub(r"<[^>]+>", "", ep.get("description_original", ""))[:300]
                if not title:
                    continue
                articles.append({
                    "title"     : title,
                    "snippet"   : snippet,
                    "date"      : "",
                    "source"    : podcast_name,
                    "media_type": "Podcast",
                    "link"      : ep.get("listennotes_url", ""),
                })
        except Exception as e:
            print(f"[Podcast] Error: {e}")

    # Deduplikasi
    seen   = set()
    unique = []
    for a in articles:
        key = a["title"][:60].lower()
        if key not in seen:
            seen.add(key)
            unique.append(a)

    print(f"[Podcast] Total episode unik: {len(unique)}")
    return unique


# ─────────────────────────────────────────────
# 4. ANALISIS SENTIMEN DENGAN CLAUDE AI
# ─────────────────────────────────────────────
def _analyze_batch(client, batch):
    """Analisis sentimen untuk satu batch artikel (maks 30)."""
    articles_text = ""
    for i, a in enumerate(batch, 1):
        articles_text += (
            f"{i}. JUDUL: {a['title']}\n"
            f"   SNIPPET: {a['snippet'][:150]}\n"
            f"   SUMBER: {a['source']}\n\n"
        )

    prompt = f"""Kamu adalah analis media monitoring profesional untuk Bank Mandiri (BMRI) Indonesia.

Berikut daftar artikel/berita:

{articles_text}

Tugasmu:
1. Tentukan sentimen: "positif", "negatif", atau "netral" terhadap citra/kinerja Bank Mandiri.
2. Tandai "irrelevant" jika tidak terkait langsung Bank Mandiri (spam, iklan, tidak ada hubungan).
3. Hitung SKOR FINAL 1–10 berdasarkan 4 dimensi berikut:

   A. DAMPAK FINANSIAL (bobot tinggi):
      - Nilai > Rp100 M              → +9–10
      - Nilai Rp10 M – Rp100 M      → +7–8
      - Nilai < Rp10 M               → +5–6
      - Tidak ada nilai rupiah        → +1–4

   B. DAMPAK REPUTASI (bobot tinggi):
      - Fraud / korupsi / skandal    → +9–10
      - Keluhan viral / regulasi     → +6–8
      - Penghargaan / milestone      → +8–9
      - Pemberitaan umum             → +3–5

   C. URGENSI & ESKALASI (bobot sedang):
      - Sudah ditangani hukum / OJK  → +8–10
      - Sedang berkembang / viral    → +6–8
      - Isu yang sudah selesai       → +3–5
      - Berita rutin / informatif    → +1–3

   D. JANGKAUAN MEDIA (bobot sedang):
      - Multi-platform + viral       → +8–10
      - Beberapa media nasional      → +6–7
      - 1–2 media saja               → +3–5
      - Media lokal / blog           → +1–3

   Skor Final = rata-rata berbobot keempat dimensi, dinormalisasi ke 1–10.
   Bulatkan ke bilangan bulat. Min 1, maks 10.

4. Berikan alasan singkat (maks 15 kata) menjelaskan skor.
5. Tentukan topik: "Keuangan", "Kasus Hukum", "Produk & Layanan", "SDM", "CSR", "Pasar Modal", "Regulasi", "Operasional", atau "Lainnya".

Output HANYA JSON array, tanpa teks lain:
[
  {{"id": 1, "sentiment": "negatif", "score": 9, "topic": "Kasus Hukum", "reason": "Alasan singkat"}},
  {{"id": 2, "sentiment": "positif", "score": 8, "topic": "Keuangan", "reason": "Alasan singkat"}},
  {{"id": 3, "sentiment": "netral",  "score": 3, "topic": "Operasional", "reason": "Alasan singkat"}},
  {{"id": 4, "sentiment": "irrelevant"}}
]"""

    try:
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = response.content[0].text.strip()

        # Ekstrak JSON array dari response
        json_match = re.search(r"\[.*\]", raw, re.DOTALL)
        if not json_match:
            print(f"[AI] Tidak ada JSON dalam response, pakai default")
            raise ValueError("No JSON found")

        scores     = json.loads(json_match.group())
        score_map  = {item["id"]: item for item in scores}

        result = []
        for i, a in enumerate(batch, 1):
            info = score_map.get(i, {})
            if info.get("sentiment") == "irrelevant":
                continue
            a["sentiment"]       = info.get("sentiment", "netral")
            a["sentiment_label"] = info.get("sentiment", "netral")
            a["score"]           = max(1, min(10, int(info.get("score", 5))))
            a["topic"]           = info.get("topic", "Lainnya")
            a["reason"]          = info.get("reason", "")
            result.append(a)
        return result

    except Exception as e:
        print(f"[AI] Error batch: {e} — pakai sentimen default")
        for a in batch:
            a.setdefault("sentiment", "netral")
            a.setdefault("score", 5)
            a.setdefault("reason", "")
        return batch


def analyze_sentiment(articles):
    """
    Kirim daftar artikel ke Claude untuk scoring sentimen (diproses per batch).
    Return list dengan field tambahan: sentiment ('positif'/'negatif'), score (1-10), reason
    """
    if not articles:
        return []

    client      = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
    BATCH_SIZE  = 30
    all_results = []
    total_batch = (len(articles) + BATCH_SIZE - 1) // BATCH_SIZE

    print(f"[AI] {len(articles)} artikel → {total_batch} batch @{BATCH_SIZE}")

    for b in range(total_batch):
        start = b * BATCH_SIZE
        batch = articles[start:start + BATCH_SIZE]
        print(f"[AI] Batch {b+1}/{total_batch} ({len(batch)} artikel)...")
        all_results.extend(_analyze_batch(client, batch))

    print(f"[AI] Selesai: {len(all_results)} artikel relevan")
    return all_results


# ─────────────────────────────────────────────
# 4b. RINGKASAN NARATIF EKSEKUTIF
# ─────────────────────────────────────────────
def generate_narrative_summary(articles, date_str):
    """
    Generate ringkasan naratif eksekutif 3 paragraf menggunakan Claude AI.
    Return string teks narasi, atau None jika gagal.
    """
    if not articles:
        return None

    client  = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
    negatif = sorted([a for a in articles if a.get("sentiment") == "negatif"],
                     key=lambda x: -x.get("score", 0))
    positif = sorted([a for a in articles if a.get("sentiment") == "positif"],
                     key=lambda x: -x.get("score", 0))

    # Rangkuman top artikel untuk prompt
    neg_list = "\n".join(
        f"- [{a['score']}/10] {a['title']} | {a['source']} | {a.get('reason','')}"
        for a in negatif[:8]
    ) or "Tidak ada"
    pos_list = "\n".join(
        f"- [{a['score']}/10] {a['title']} | {a['source']} | {a.get('reason','')}"
        for a in positif[:8]
    ) or "Tidak ada"

    # Hitung breakdown per kanal
    from collections import Counter
    kanal = Counter(a.get("media_type", "News") for a in articles)
    kanal_str = ", ".join(f"{k}: {v}" for k, v in kanal.items())

    prompt = f"""Kamu adalah analis komunikasi senior Bank Mandiri.

DATA MEDIA MONITORING — {date_str}
Total artikel: {len(articles)} | Negatif: {len(negatif)} | Positif: {len(positif)}
Kanal: {kanal_str}

TOP BERITA NEGATIF:
{neg_list}

TOP BERITA POSITIF:
{pos_list}

Tulis RINGKASAN NARATIF EKSEKUTIF dalam Bahasa Indonesia yang formal dan padat.
Format TEPAT 3 paragraf:
1. Gambaran umum sentimen media hari ini dan tema dominan
2. Isu-isu negatif utama yang perlu perhatian manajemen (jika ada, sebutkan konkret)
3. Poin positif menonjol dan rekomendasi tindak lanjut singkat

Maksimal 200 kata total. Langsung ke inti, tanpa pembuka seperti "Berikut ringkasan..."."""

    try:
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=512,
            messages=[{"role": "user", "content": prompt}]
        )
        narasi = response.content[0].text.strip()
        print("[Narasi] Ringkasan naratif berhasil dibuat.")
        return narasi
    except Exception as e:
        print(f"[Narasi] Error: {e}")
        return None


# ─────────────────────────────────────────────
# 5. BUAT GRAFIK SENTIMEN
# ─────────────────────────────────────────────
def generate_chart(articles, output_path):
    """
    Grafik dark-theme 3-kolom:
      Kiri (22%)   : Donut sentimen + tabel distribusi media
      Tengah (37%) : Top artikel NEGATIF (judul di atas bar)
      Kanan (37%)  : Top artikel POSITIF (judul di atas bar)
    Menggunakan fig.add_axes() untuk posisi presisi tanpa overlap GridSpec.
    """
    BG      = "#0D1117"
    CARD_BG = "#161B22"
    NEG_C   = "#F85149"
    POS_C   = "#3FB950"
    TEXT_C  = "#E6EDF3"
    MUTED_C = "#8B949E"
    GRID_C  = "#21262D"

    negatif_all = sorted(
        [a for a in articles if a.get("sentiment") == "negatif"],
        key=lambda x: x.get("score", 0), reverse=True,
    )
    positif_all = sorted(
        [a for a in articles if a.get("sentiment") == "positif"],
        key=lambda x: x.get("score", 0), reverse=True,
    )
    total   = len(articles)
    n_neg   = len(negatif_all)
    n_pos   = len(positif_all)
    pct_pos = round(n_pos / total * 100) if total else 0
    pct_neg = 100 - pct_pos

    negatif  = negatif_all[:10]
    positif  = positif_all[:10]
    yesterday = (datetime.date.today() - datetime.timedelta(days=1)).strftime("%d %B %Y")

    # Tinggi gambar: setiap artikel butuh 2 baris (judul + bar), min 10"
    n_rows = max(len(negatif), len(positif), 5)
    fig_h  = max(10, n_rows * 0.95 + 4.5)

    fig = plt.figure(figsize=(26, fig_h), facecolor=BG)

    # ── Judul ────────────────────────────────────────────
    fig.text(0.5, 0.975,
             f"Media Monitoring Bank Mandiri (BMRI)  —  {yesterday}",
             ha="center", fontsize=16, fontweight="bold", color=TEXT_C)
    fig.text(0.5, 0.958,
             f"Total {total} artikel dianalisis oleh Claude AI",
             ha="center", fontsize=9.5, color=MUTED_C)

    # ── Posisi panel (figure fraction: left, bottom, width, height) ──────
    # Kiri: x=0.02..0.23  (donut atas, tabel media bawah)
    # Neg : x=0.26..0.61
    # Pos : x=0.64..0.97
    BODY_B = 0.07    # bottom padding
    BODY_T = 0.93    # top (di bawah judul)
    BODY_H = BODY_T - BODY_B

    # Donut: 55% atas dari panel kiri
    ax_donut = fig.add_axes([0.02, BODY_B + BODY_H * 0.42, 0.21, BODY_H * 0.54])
    ax_donut.set_facecolor(CARD_BG)

    # Tabel media: 40% bawah dari panel kiri
    ax_tbl = fig.add_axes([0.02, BODY_B, 0.21, BODY_H * 0.40])
    ax_tbl.set_facecolor(CARD_BG)
    ax_tbl.axis("off")

    # Negatif & Positif bar panels
    ax_neg = fig.add_axes([0.26, BODY_B, 0.35, BODY_H])
    ax_neg.set_facecolor(CARD_BG)
    ax_pos = fig.add_axes([0.63, BODY_B, 0.35, BODY_H])
    ax_pos.set_facecolor(CARD_BG)

    # ════════════════════════════════════════════════════
    # DONUT
    # ════════════════════════════════════════════════════
    overall_label = "POSITIF ✓" if pct_pos >= 60 else ("NEGATIF !" if pct_neg > 60 else "NETRAL ~")
    overall_color = POS_C if pct_pos >= 60 else (NEG_C if pct_neg > 60 else "#F0A500")
    ax_donut.pie(
        [n_neg, n_pos] if total else [1, 1],
        colors=[NEG_C, POS_C],
        startangle=90, counterclock=False,
        wedgeprops=dict(width=0.38, edgecolor=CARD_BG, linewidth=3),
    )
    ax_donut.text(0,  0.10, f"{pct_pos}%", ha="center", va="center",
                  fontsize=28, fontweight="bold", color=overall_color)
    ax_donut.text(0, -0.24, overall_label, ha="center", va="center",
                  fontsize=10, fontweight="bold", color=overall_color)
    ax_donut.set_title("SENTIMEN KESELURUHAN", fontsize=9, color=MUTED_C,
                       fontweight="bold", pad=10)
    ax_donut.legend(
        [f"Negatif  {n_neg} artikel ({pct_neg}%)",
         f"Positif  {n_pos} artikel ({pct_pos}%)"],
        loc="lower center", bbox_to_anchor=(0.5, -0.10),
        fontsize=9, framealpha=0, labelcolor=[NEG_C, POS_C],
    )

    # ════════════════════════════════════════════════════
    # TABEL MEDIA
    # ════════════════════════════════════════════════════
    media_types = ["News", "Blog", "Twitter", "Google Alerts", "Podcast", "E-Commerce"]
    MTYPE_COLOR = {
        "News": "#58A6FF", "Blog": "#3FB950", "Twitter": "#1D9BF0",
        "Google Alerts": "#F0A500", "Podcast": "#BC8CFF", "E-Commerce": "#FF7B72",
    }
    rows_data = []
    for mt in media_types:
        sub = [a for a in articles if a.get("media_type") == mt]
        if not sub:
            continue
        neg_c = sum(1 for a in sub if a.get("sentiment") == "negatif")
        rows_data.append((mt, len(sub), neg_c, len(sub) - neg_c))

    ax_tbl.set_title("DISTRIBUSI PER TIPE MEDIA", fontsize=9, color=MUTED_C,
                     fontweight="bold", pad=8)
    col_x = [0.03, 0.60, 0.74, 0.89]
    for xi, h in zip(col_x, ["Tipe Media", "Total", "Neg", "Pos"]):
        ax_tbl.text(xi, 0.97, h, transform=ax_tbl.transAxes,
                    fontsize=8.5, color=MUTED_C, fontweight="bold", va="top")
    ax_tbl.plot([0, 1], [0.90, 0.90], transform=ax_tbl.transAxes,
                color=GRID_C, linewidth=0.8, clip_on=False)

    step = 0.84 / max(len(rows_data), 1)
    for r_i, (mt, tot, neg_c, pos_c) in enumerate(rows_data):
        y = 0.87 - r_i * step
        c = MTYPE_COLOR.get(mt, TEXT_C)
        ax_tbl.text(col_x[0], y, mt,         transform=ax_tbl.transAxes,
                    fontsize=9, color=c, va="center")
        ax_tbl.text(col_x[1], y, str(tot),   transform=ax_tbl.transAxes,
                    fontsize=9, color=TEXT_C, va="center", ha="center")
        ax_tbl.text(col_x[2], y, str(neg_c), transform=ax_tbl.transAxes,
                    fontsize=9, color=NEG_C, va="center", ha="center")
        ax_tbl.text(col_x[3], y, str(pos_c), transform=ax_tbl.transAxes,
                    fontsize=9, color=POS_C, va="center", ha="center")

    # ════════════════════════════════════════════════════
    # HELPER: Panel artikel (judul DI ATAS bar, skor di ujung bar)
    # ════════════════════════════════════════════════════
    def make_article_panel(ax, items, color, header, n_total):
        ax.spines[:].set_visible(False)
        ax.set_xlim(0, 11.5)
        # Setiap artikel = 2 unit: 0–1 = judul+sumber, 1–2 = bar
        n = len(items)
        ax.set_ylim(0, n * 2 + 0.4)
        ax.invert_yaxis()   # artikel teratas (skor tinggi) di atas

        ax.set_title(f"{header}   ({n_total} artikel)",
                     fontsize=11, color=color, fontweight="bold", pad=12)

        for i, a in enumerate(items):
            score = a.get("score", 0)
            title = a["title"]
            title_disp = title[:54] + "…" if len(title) > 54 else title
            src   = a.get("source", "")[:22]

            y_title = i * 2 + 0.45   # pusat baris judul
            y_src   = i * 2 + 0.90   # pusat baris sumber
            y_bar   = i * 2 + 1.45   # pusat bar

            # Judul artikel
            ax.text(0.15, y_title, title_disp,
                    va="center", ha="left", fontsize=8.5,
                    color=TEXT_C, clip_on=True)
            # Sumber
            ax.text(0.15, y_src, f"[{src}]",
                    va="center", ha="left", fontsize=7.5,
                    color=MUTED_C, clip_on=True)
            # Bar
            ax.barh(y_bar, score, height=0.45,
                    color=color, alpha=0.88, zorder=3)
            # Skor di ujung bar (tidak tumpuk karena bar dan judul di baris berbeda)
            ax.text(score + 0.25, y_bar, str(score),
                    va="center", ha="left", fontsize=10.5,
                    color=color, fontweight="bold")
            # Separator antar artikel
            if i < n - 1:
                ax.axhline(y=i * 2 + 2.0, color=GRID_C,
                           linewidth=0.6, zorder=1)

        # x-axis
        ax.tick_params(left=False, labelleft=False)
        ax.set_xticks([2, 4, 6, 8, 10])
        ax.tick_params(axis="x", colors=MUTED_C, labelsize=8.5, length=4)
        ax.set_xlabel("→  Skor Sentimen", fontsize=9, color=MUTED_C, labelpad=8)
        ax.grid(axis="x", color=GRID_C, linewidth=0.5, zorder=0)

    make_article_panel(ax_neg, negatif, NEG_C, "▼  NEGATIF", n_neg)
    make_article_panel(ax_pos, positif, POS_C, "▲  POSITIF", n_pos)

    # ── Footer ───────────────────────────────────────────
    fig.text(
        0.5, 0.028,
        "Analisis otomatis oleh Claude AI (Anthropic)  ·  "
        "Sumber: Talkwalker Alerts · Google Alerts · ListenNotes",
        ha="center", fontsize=8.5, color=MUTED_C,
    )

    plt.savefig(output_path, dpi=150, bbox_inches="tight", facecolor=BG)
    plt.close()
    print(f"[Chart] Grafik disimpan: {output_path}")


# ─────────────────────────────────────────────
# 6. FORMAT KONTEN TELEGRAM & EMAIL
# ─────────────────────────────────────────────
def format_telegram_negative(articles, date_str):
    negatif = sorted(
        [a for a in articles if a.get("sentiment") == "negatif"],
        key=lambda x: x["score"], reverse=True
    )
    if not negatif:
        return "✅ Tidak ada berita negatif hari ini."

    lines = [
        f"🔴 *MEDIA MONITORING BANK MANDIRI — NEGATIF*",
        f"📅 {date_str}",
        f"",
        f"Total: *{len(negatif)} berita negatif*",
        f"─────────────────────",
    ]
    for i, a in enumerate(negatif, 1):
        lines += [
            f"",
            f"*{i}. Skor: {a['score']}/10*",
            f"📰 {a['title'][:120]}",
            f"🔍 _{a['reason']}_",
            f"📡 `{a['source']}`  |  {a['date']}",
        ]
    lines += ["", "─────────────────────",
              "_Analisis otomatis oleh Claude AI_"]
    return "\n".join(lines)


def format_telegram_positive(articles, date_str):
    positif = sorted(
        [a for a in articles if a.get("sentiment") == "positif"],
        key=lambda x: x["score"], reverse=True
    )
    if not positif:
        return "ℹ️ Tidak ada berita positif signifikan hari ini."

    lines = [
        f"🟢 *MEDIA MONITORING BANK MANDIRI — POSITIF*",
        f"📅 {date_str}",
        f"",
        f"Total: *{len(positif)} berita positif*",
        f"─────────────────────",
    ]
    for i, a in enumerate(positif, 1):
        lines += [
            f"",
            f"*{i}. Skor: {a['score']}/10*",
            f"📰 {a['title'][:120]}",
            f"💡 _{a['reason']}_",
            f"📡 `{a['source']}`  |  {a['date']}",
        ]
    lines += ["", "─────────────────────",
              "_Analisis otomatis oleh Claude AI_"]
    return "\n".join(lines)


def format_summary(articles, date_str):
    positif = [a for a in articles if a.get("sentiment") == "positif"]
    negatif = [a for a in articles if a.get("sentiment") == "negatif"]
    total   = len(articles)
    pct_pos = round(len(positif) / total * 100) if total else 0

    avg_pos = round(sum(a["score"] for a in positif) / len(positif), 1) if positif else 0
    avg_neg = round(sum(a["score"] for a in negatif) / len(negatif), 1) if negatif else 0
    top_neg = negatif[0] if negatif else None
    top_pos = positif[0] if positif else None

    sentiment_icon = "🟢" if pct_pos >= 60 else ("🟡" if pct_pos >= 40 else "🔴")

    lines = [
        f"📊 *RINGKASAN MEDIA MONITORING BANK MANDIRI*",
        f"📅 {date_str}",
        f"",
        f"{sentiment_icon} Sentimen Keseluruhan: *{'POSITIF' if pct_pos >= 60 else 'NEGATIF' if pct_pos < 40 else 'NETRAL'}* ({pct_pos}%)",
        f"",
        f"📈 Total Berita Relevan : {total}",
        f"🟢 Positif  : {len(positif)} berita  |  Avg Skor: {avg_pos}",
        f"🔴 Negatif  : {len(negatif)} berita  |  Avg Skor: {avg_neg}",
    ]

    if top_neg:
        lines += [
            f"",
            f"⚠️ *Berita Negatif Utama:*",
            f"_{top_neg['title'][:100]}_",
            f"Skor {top_neg['score']}/10 — {top_neg['source']}",
        ]
    if top_pos:
        lines += [
            f"",
            f"✨ *Berita Positif Utama:*",
            f"_{top_pos['title'][:100]}_",
            f"Skor {top_pos['score']}/10 — {top_pos['source']}",
        ]

    lines += ["", "_📬 Laporan dikirim ke Email & Telegram_",
              "_Analisis otomatis oleh Claude AI_"]
    return "\n".join(lines)


# ─────────────────────────────────────────────
# 7. KIRIM KE TELEGRAM
# ─────────────────────────────────────────────
def send_telegram_text(text):
    url  = f"https://api.telegram.org/bot{TG_TOKEN}/sendMessage"
    data = {
        "chat_id"    : TG_CHAT_ID,
        "text"       : text,
        "parse_mode" : "Markdown",
    }
    r = requests.post(url, data=data, timeout=30)
    if not r.ok:
        print(f"[Telegram] Error text: {r.text}")
    else:
        print("[Telegram] Teks terkirim.")


def send_telegram_photo(image_path, caption=""):
    url  = f"https://api.telegram.org/bot{TG_TOKEN}/sendPhoto"
    with open(image_path, "rb") as f:
        r = requests.post(
            url,
            data={"chat_id": TG_CHAT_ID, "caption": caption, "parse_mode": "Markdown"},
            files={"photo": f},
            timeout=60,
        )
    if not r.ok:
        print(f"[Telegram] Error photo: {r.text}")
    else:
        print("[Telegram] Grafik terkirim.")


# ─────────────────────────────────────────────
# 8. KIRIM EMAIL (SMTP GMAIL)
# ─────────────────────────────────────────────
def build_html_email(articles, date_str, chart_path, narrative=None):
    """Buat HTML email lengkap dengan tabel berita dan grafik embedded."""
    positif = sorted(
        [a for a in articles if a.get("sentiment") == "positif"],
        key=lambda x: x["score"], reverse=True
    )
    negatif = sorted(
        [a for a in articles if a.get("sentiment") == "negatif"],
        key=lambda x: x["score"], reverse=True
    )
    total   = len(articles)
    pct_pos = round(len(positif) / total * 100) if total else 0

    def rows(items, color, badge_color):
        html = ""
        for i, a in enumerate(items, 1):
            bar_width = int(a["score"] * 10)
            link = a.get("link", "").strip()
            if link:
                title_cell = (
                    f'<a href="{link}" target="_blank" rel="noopener" '
                    f'style="font-weight:600;color:#1d4ed8;font-size:13px;line-height:1.4;'
                    f'text-decoration:none">{a["title"]} ↗</a>'
                )
            else:
                title_cell = (
                    f'<span style="font-weight:600;color:#1a1a1a;font-size:13px;'
                    f'line-height:1.4">{a["title"]}</span>'
                )
            html += f"""
            <tr style="border-bottom:1px solid #f0f0f0">
              <td style="padding:10px 8px;text-align:center;font-weight:700;color:{color};font-size:15px">{a['score']}</td>
              <td style="padding:10px 8px">
                <div>{title_cell}</div>
                <div style="color:#666;font-size:11px;margin-top:4px">{a.get('reason','')}</div>
                <div style="background:#eee;border-radius:3px;height:4px;margin-top:6px;overflow:hidden">
                  <div style="background:{color};width:{bar_width}%;height:4px;border-radius:3px"></div>
                </div>
              </td>
              <td style="padding:10px 8px;font-size:11px;color:#888;white-space:nowrap">
                <span style="background:{badge_color};color:{color};padding:2px 7px;border-radius:10px;font-size:10px;font-weight:600">{a.get('media_type','')}</span><br>
                <span style="margin-top:4px;display:block">{a['source']}</span>
                <span>{a.get('date','')}</span>
              </td>
            </tr>"""
        return html

    neg_rows = rows(negatif, "#C0392B", "#FDECEA")
    pos_rows = rows(positif, "#27AE60", "#E8F5E9")

    sentiment_label = "POSITIF ✅" if pct_pos >= 60 else ("NEGATIF ⚠️" if pct_pos < 40 else "NETRAL ⚡")
    sentiment_color = "#27AE60" if pct_pos >= 60 else ("#C0392B" if pct_pos < 40 else "#F39C12")

    html = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Media Monitoring Bank Mandiri — {date_str}</title></head>
<body style="margin:0;padding:0;background:#F5F6FA;font-family:Arial,Helvetica,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#F5F6FA;padding:20px 0">
<tr><td align="center">
<table width="680" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.08)">

  <!-- Header -->
  <tr><td style="background:linear-gradient(135deg,#003087,#0055B3);padding:28px 32px">
    <div style="color:#fff;font-size:22px;font-weight:700">📊 Media Monitoring Bank Mandiri</div>
    <div style="color:#B0C8FF;font-size:13px;margin-top:6px">{date_str} &nbsp;|&nbsp; Sumber: Talkwalker Alerts &nbsp;|&nbsp; Analisis: Claude AI</div>
  </td></tr>

  <!-- Summary Cards -->
  <tr><td style="padding:24px 32px 16px">
    <table width="100%" cellpadding="0" cellspacing="0">
      <tr>
        <td width="25%" style="padding:4px">
          <div style="background:#F0F4FF;border-radius:10px;padding:14px 12px;text-align:center">
            <div style="font-size:28px;font-weight:700;color:#003087">{total}</div>
            <div style="font-size:11px;color:#666;margin-top:4px">Total Berita</div>
          </div>
        </td>
        <td width="25%" style="padding:4px">
          <div style="background:#FEF0F0;border-radius:10px;padding:14px 12px;text-align:center">
            <div style="font-size:28px;font-weight:700;color:#C0392B">{len(negatif)}</div>
            <div style="font-size:11px;color:#666;margin-top:4px">Berita Negatif</div>
          </div>
        </td>
        <td width="25%" style="padding:4px">
          <div style="background:#F0FBF4;border-radius:10px;padding:14px 12px;text-align:center">
            <div style="font-size:28px;font-weight:700;color:#27AE60">{len(positif)}</div>
            <div style="font-size:11px;color:#666;margin-top:4px">Berita Positif</div>
          </div>
        </td>
        <td width="25%" style="padding:4px">
          <div style="background:#FFF8E6;border-radius:10px;padding:14px 12px;text-align:center">
            <div style="font-size:22px;font-weight:700;color:{sentiment_color}">{pct_pos}%</div>
            <div style="font-size:11px;color:{sentiment_color};margin-top:4px;font-weight:600">{sentiment_label}</div>
          </div>
        </td>
      </tr>
    </table>
  </td></tr>

  <!-- Chart -->
  <tr><td style="padding:8px 32px 8px">
    <div style="font-size:14px;font-weight:700;color:#333;margin-bottom:10px">📈 Grafik Sentimen</div>
    <img src="cid:sentiment_chart" width="100%" style="border-radius:8px;border:1px solid #eee" alt="Grafik Sentimen">
  </td></tr>

  <!-- Negative News -->
  <tr><td style="padding:20px 32px 8px">
    <div style="font-size:14px;font-weight:700;color:#C0392B;margin-bottom:10px">🔴 Berita Negatif — Sorted by Score</div>
    <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #F5E6E6;border-radius:8px;overflow:hidden">
      <tr style="background:#FEF0F0">
        <th style="padding:8px;text-align:center;font-size:11px;color:#888;width:50px">SKOR</th>
        <th style="padding:8px;text-align:left;font-size:11px;color:#888">JUDUL & ANALISIS</th>
        <th style="padding:8px;text-align:left;font-size:11px;color:#888;width:130px">SUMBER</th>
      </tr>
      {neg_rows if neg_rows else '<tr><td colspan="3" style="padding:14px;text-align:center;color:#888;font-size:12px">Tidak ada berita negatif hari ini ✅</td></tr>'}
    </table>
  </td></tr>

  <!-- Positive News -->
  <tr><td style="padding:20px 32px 8px">
    <div style="font-size:14px;font-weight:700;color:#27AE60;margin-bottom:10px">🟢 Berita Positif — Sorted by Score</div>
    <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #E6F5EA;border-radius:8px;overflow:hidden">
      <tr style="background:#F0FBF4">
        <th style="padding:8px;text-align:center;font-size:11px;color:#888;width:50px">SKOR</th>
        <th style="padding:8px;text-align:left;font-size:11px;color:#888">JUDUL & ANALISIS</th>
        <th style="padding:8px;text-align:left;font-size:11px;color:#888;width:130px">SUMBER</th>
      </tr>
      {pos_rows if pos_rows else '<tr><td colspan="3" style="padding:14px;text-align:center;color:#888;font-size:12px">Tidak ada berita positif hari ini.</td></tr>'}
    </table>
  </td></tr>

  <!-- Narrative Summary -->
  {f'''<tr><td style="padding:20px 32px 8px">
    <div style="font-size:14px;font-weight:700;color:#1F4E79;margin-bottom:10px">📋 Ringkasan Naratif Eksekutif</div>
    <div style="background:#F0F4FF;border-left:4px solid #2E75B6;border-radius:0 8px 8px 0;padding:16px 18px;font-size:13px;line-height:1.7;color:#1a1a1a">
      {narrative.replace(chr(10), '<br>')}
    </div>
  </td></tr>''' if narrative else ''}

  <!-- Footer -->
  <tr><td style="padding:20px 32px;background:#F8F9FB;border-top:1px solid #eee;text-align:center">
    <div style="font-size:11px;color:#999">
      Laporan ini dibuat secara otomatis oleh sistem Media Monitoring Bank Mandiri<br>
      Powered by <strong>Claude AI (Anthropic)</strong> &amp; Talkwalker Alerts<br>
      Dikirim setiap hari pukul 07.00 WIB
    </div>
  </td></tr>

</table>
</td></tr>
</table>
</body>
</html>"""
    return html


def generate_excel_report(articles, date_str):
    """
    Buat file Excel laporan media monitoring.
    Sheet 1: Laporan Harian (negatif → positif, sorted by score)
    Sheet 2: Daftar Media (semua sumber yang diproses)
    Return path ke file .xlsx sementara.
    """
    from collections import Counter

    wb = openpyxl.Workbook()

    # ── Helper styles ──────────────────────────────────────────────
    def border():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    def hfont(color="FFFFFF", sz=10, bold=True):
        return Font(name="Arial", bold=bold, color=color, size=sz)

    def cfont(bold=False, sz=10):
        return Font(name="Arial", bold=bold, size=sz)

    def fill(hex_color):
        return PatternFill("solid", start_color=hex_color)

    def center(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    def left(indent=1, wrap=False):
        return Alignment(horizontal="left", vertical="center", indent=indent, wrap_text=wrap)

    negatif = sorted([a for a in articles if a.get("sentiment") == "negatif"],
                     key=lambda x: -x.get("score", 0))
    positif = sorted([a for a in articles if a.get("sentiment") == "positif"],
                     key=lambda x: -x.get("score", 0))
    total   = len(articles)
    pct_pos = round(len(positif) / total * 100) if total else 0

    # ═══════════════════════════════════════════════════════════════
    # SHEET 1 — LAPORAN HARIAN
    # ═══════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Laporan Harian"
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 38

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = f"MEDIA MONITORING BANK MANDIRI — {date_str}"
    ws["A1"].font = hfont(sz=13)
    ws["A1"].fill = fill("1F4E79")
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    # Sub-title
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Total: {total} artikel  |  🔴 Negatif: {len(negatif)}  |  🟢 Positif: {len(positif)}  |  Sentimen Positif: {pct_pos}%"
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="444444")
    ws["A2"].fill = fill("D6E4F0")
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 20

    row = 4

    for section, data, hdr_color, row_color_odd, row_color_even, label in [
        ("🔴 BERITA NEGATIF", negatif, "C0392B", "FEF0F0", "FDDDDD", "NEGATIF"),
        ("🟢 BERITA POSITIF", positif, "27AE60", "F0FBF4", "DCFBE8", "POSITIF"),
    ]:
        # Section header
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = f"{section}  —  {len(data)} artikel"
        ws[f"A{row}"].font = hfont(sz=11)
        ws[f"A{row}"].fill = fill(hdr_color)
        ws[f"A{row}"].alignment = left()
        ws.row_dimensions[row].height = 24
        row += 1

        # Column headers
        col_hdrs = ["No", "Judul Artikel", "Sumber Media", "Skor", "Analisis AI"]
        for c, h in enumerate(col_hdrs, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = hfont(sz=10)
            cell.fill = fill("2E75B6")
            cell.alignment = center()
            cell.border = border()
        ws.row_dimensions[row].height = 22
        row += 1

        if not data:
            ws.merge_cells(f"A{row}:E{row}")
            ws[f"A{row}"] = "Tidak ada berita dalam kategori ini."
            ws[f"A{row}"].font = Font(name="Arial", italic=True, size=10, color="888888")
            ws[f"A{row}"].alignment = center()
            ws.row_dimensions[row].height = 20
            row += 2
            continue

        for i, art in enumerate(data, 1):
            bg = row_color_odd if i % 2 == 1 else row_color_even
            vals = [
                i,
                art.get("title", ""),
                art.get("source", ""),
                art.get("score", ""),
                art.get("reason", ""),
            ]
            aligns = [center(), left(wrap=True), left(), center(), left(wrap=True)]
            for c, (v, al) in enumerate(zip(vals, aligns), 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = cfont(sz=10)
                cell.fill = fill(bg)
                cell.alignment = al
                cell.border = border()
            ws.row_dimensions[row].height = 32
            row += 1

        row += 1  # spacer between sections

    # ═══════════════════════════════════════════════════════════════
    # SHEET 2 — DAFTAR MEDIA
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Daftar Media")
    ws2.column_dimensions["A"].width = 7
    ws2.column_dimensions["B"].width = 42
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 26

    ws2.merge_cells("A1:E1")
    ws2["A1"] = f"DAFTAR MEDIA YANG DIPROSES — {date_str}"
    ws2["A1"].font = hfont(sz=13)
    ws2["A1"].fill = fill("1F4E79")
    ws2["A1"].alignment = center()
    ws2.row_dimensions[1].height = 30

    # Hitung sumber per media_type dari articles
    source_by_type = {}
    for art in articles:
        mt  = art.get("media_type", "News")
        src = art.get("source", "")
        if src:
            source_by_type.setdefault(mt, Counter())[src] += 1

    row2 = 3
    no   = 1
    TYPE_META = [
        ("News",          "📰", "1F4E79", "D6E4F0"),
        ("Blog",          "📝", "375623", "E2EFDA"),
        ("Twitter",       "🐦", "1C3557", "DAE3F3"),
        ("Google Alerts", "🔍", "B45309", "FEF3C7"),
        ("Podcast",       "🎙️", "6B21A8", "F3E8FF"),
        ("E-Commerce",    "🛒", "065F46", "D1FAE5"),
    ]

    for mt, emoji, hdr_hex, row_hex in TYPE_META:
        counter = source_by_type.get(mt, Counter())
        if not counter:
            continue

        # Section header
        ws2.merge_cells(f"A{row2}:E{row2}")
        ws2[f"A{row2}"] = f"{emoji} {mt.upper()}  —  {sum(counter.values())} artikel  |  {len(counter)} sumber"
        ws2[f"A{row2}"].font = hfont(sz=10)
        ws2[f"A{row2}"].fill = fill(hdr_hex)
        ws2[f"A{row2}"].alignment = left()
        ws2.row_dimensions[row2].height = 22
        row2 += 1

        # Column headers
        for c, h in enumerate(["No", "Nama Media / Sumber", "Tipe", "Jml Artikel", "Status"], 1):
            cell = ws2.cell(row=row2, column=c, value=h)
            cell.font = hfont(sz=10)
            cell.fill = fill("2E75B6")
            cell.alignment = center()
            cell.border = border()
        ws2.row_dimensions[row2].height = 22
        row2 += 1

        for i, (src, cnt) in enumerate(sorted(counter.items()), 1):
            bg = row_hex if i % 2 == 0 else "FFFFFF"
            vals = [no, src, mt, cnt, "✅ Berhasil Diproses"]
            alns = [center(), left(), center(), center(), center()]
            for c, (v, al) in enumerate(zip(vals, alns), 1):
                cell = ws2.cell(row=row2, column=c, value=v)
                cell.font = cfont(sz=10)
                cell.fill = fill(bg)
                cell.alignment = al
                cell.border = border()
            ws2.row_dimensions[row2].height = 18
            no   += 1
            row2 += 1

        row2 += 1  # spacer

    # Freeze header rows
    ws["A3"].offset  # trigger
    ws.freeze_panes  = "A4"
    ws2.freeze_panes = "A3"

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)
    return tmp.name


# ─────────────────────────────────────────────
# 8b. SAVE ARTICLES HISTORY (docs/articles.json)
# ─────────────────────────────────────────────
def save_articles_history(articles, date_str, narrative, docs_dir):
    """Simpan artikel harian ke docs/articles.json (akumulatif, max 1 tahun)."""
    import json as _json
    from datetime import datetime as _dt, timedelta as _td

    os.makedirs(docs_dir, exist_ok=True)
    json_path = os.path.join(docs_dir, "articles.json")

    run_date = None
    for fmt in ("%d %B %Y %H:%M WIB", "%d %B %Y"):
        try:
            run_date = _dt.strptime(date_str.strip(), fmt).strftime("%Y-%m-%d")
            break
        except Exception:
            continue
    if not run_date:
        # Fallback: UTC+7
        run_date = (_dt.utcnow() + _td(hours=7)).strftime("%Y-%m-%d")

    history = []
    if os.path.exists(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                history = _json.load(f)
        except Exception:
            history = []

    # Hapus data tanggal yang sama (re-run idempoten)
    history = [a for a in history if a.get("run_date") != run_date]

    # Filter artikel tidak relevan — tidak masuk ke history
    _EXCLUDE_KW = [
        'bank syariah mandiri', 'mandiri syariah', 'bsm.co.id',  # BSM
        'hanania',  # Travel umrah Hanania — rekening tersangka di BM, bukan kasus BM
    ]
    def _is_excluded(a):
        t = (a.get("title","") + " " + a.get("snippet","") + " " + a.get("source","")).lower()
        return any(k in t for k in _EXCLUDE_KW)
    articles = [a for a in articles if not _is_excluded(a)]

    # Tambah artikel hari ini
    for a in articles:
        history.append({
            "run_date":   run_date,
            "title":      a.get("title", ""),
            "source":     a.get("source", ""),
            "sentiment":  a.get("sentiment", ""),
            "score":      a.get("score", 0),
            "reason":     a.get("reason", ""),
            "media_type": a.get("media_type", "News"),
            "link":       a.get("link", ""),
        })

    # Hapus data lebih dari 366 hari
    cutoff_date = (_dt.now() - _td(days=366)).strftime("%Y-%m-%d")
    history = [a for a in history if a.get("run_date", "") >= cutoff_date]
    history.sort(key=lambda x: x.get("run_date", ""), reverse=True)

    with open(json_path, "w", encoding="utf-8") as f:
        _json.dump(history, f, ensure_ascii=False, separators=(",", ":"))

    if narrative:
        # Simpan ke narratives.json (akumulatif per tanggal)
        nar_path = os.path.join(docs_dir, "narratives.json")
        narr_data = {}
        if os.path.exists(nar_path):
            try:
                with open(nar_path, "r", encoding="utf-8") as f:
                    narr_data = _json.load(f)
            except Exception:
                narr_data = {}
        narr_data[run_date] = narrative
        # Prune narasi lama
        narr_data = {k: v for k, v in narr_data.items() if k >= cutoff_date}
        with open(nar_path, "w", encoding="utf-8") as f:
            _json.dump(narr_data, f, ensure_ascii=False, indent=None, separators=(",", ":"))

    oldest = history[-1]["run_date"] if history else "-"
    print(f"  ✅ Historis: {len(history)} artikel tersimpan (sejak {oldest})")
    return run_date


# ─────────────────────────────────────────────
# 8c. GENERATE TV DASHBOARD (GitHub Pages)
# ─────────────────────────────────────────────
def generate_tv_dashboard(articles, date_str, chart_path=None, narrative=None):
    """
    Generate HTML dashboard TV dengan time filter (Kemarin s/d 1 Tahun).
    Load data dari docs/articles.json via JS fetch(). Render dinamis di browser.
    Return string HTML.
    """
    import json as _json

    # Tampilkan UTC dan WIB; tanggal menggunakan WIB (bisa berbeda saat UTC 17:xx → WIB 00:xx+1)
    _now_utc = datetime.datetime.utcnow()
    _now_wib = _now_utc + datetime.timedelta(hours=7)
    now_str  = (_now_wib.strftime("%d %B %Y") + ", "
                + _now_utc.strftime("%H:%M") + " UTC"
                + " (" + _now_wib.strftime("%H:%M") + " WIB)")
    run_date = _now_wib.strftime("%Y-%m-%d")

    # Embedded JSON fallback (untuk tampilan pertama sebelum articles.json dimuat)
    embedded = [
        {
            "run_date":   run_date,
            "title":      a.get("title", ""),
            "source":     a.get("source", ""),
            "sentiment":  a.get("sentiment", ""),
            "score":      a.get("score", 0),
            "reason":     a.get("reason", ""),
            "media_type": a.get("media_type", "News"),
            "link":       a.get("link", ""),
        }
        for a in articles
    ]
    embedded_json = _json.dumps(embedded, ensure_ascii=False, separators=(",", ":"))
    nar_safe = (narrative or "").replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${").replace("\n", "<br>")

    # ── CSS ──────────────────────────────────────────────────────────────────
    css = (
        "*{margin:0;padding:0;box-sizing:border-box}"
        "body{background:#080d1a;color:#e2e8f0;font-family:'Segoe UI',Arial,sans-serif;padding:16px 24px}"
        ".hdr{text-align:center;padding:12px 0 14px;border-bottom:1px solid #1e293b;margin-bottom:11px}"
        ".hdr h1{font-size:1.65rem;font-weight:800;color:#fff}"
        ".hdr .sub{color:#64748b;font-size:.78rem;margin-top:4px}"
        ".hdr .upd{color:#334155;font-size:.68rem;margin-top:3px}"
        ".filter-bar{display:flex;align-items:center;gap:8px;background:#0f172a;border:1px solid #1e293b;"
        "border-radius:10px;padding:8px 14px;margin-bottom:11px;flex-wrap:wrap}"
        ".filter-label{font-size:.68rem;font-weight:700;color:#64748b;text-transform:uppercase;"
        "letter-spacing:.6px;margin-right:4px;white-space:nowrap}"
        ".filter-btn{padding:5px 13px;border-radius:20px;font-size:.74rem;font-weight:600;"
        "border:1px solid #1e293b;background:transparent;color:#94a3b8;cursor:pointer;transition:all .15s;white-space:nowrap}"
        ".filter-btn:hover{border-color:#3b82f6;color:#93c5fd}"
        ".filter-btn.active{background:#1d4ed8;border-color:#1d4ed8;color:#fff}"
        ".filter-divider{width:1px;height:16px;background:#1e293b;margin:0 2px}"
        ".filter-info{margin-left:auto;font-size:.67rem;color:#475569;white-space:nowrap}"
        ".stats-row{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:11px}"
        ".stat-card{background:#0f172a;border:1px solid #1e293b;border-radius:10px;padding:12px 8px;text-align:center}"
        ".stat-num{font-size:1.85rem;font-weight:800;line-height:1}"
        ".stat-lbl{font-size:.6rem;color:#64748b;margin-top:4px;text-transform:uppercase;letter-spacing:.6px}"
        ".click-legend{display:flex;gap:18px;justify-content:center;align-items:center;background:#0f172a;"
        "border-radius:7px;padding:5px 14px;margin-bottom:11px;font-size:.72rem;color:#64748b}"
        ".cl-dot{width:8px;height:8px;border-radius:50%;display:inline-block;margin-right:5px;vertical-align:middle}"
        ".main-grid{display:grid;grid-template-columns:210px 1fr 1fr;gap:11px;align-items:start;margin-bottom:11px}"
        ".left-panel{background:#0f172a;border:1px solid #1e293b;border-radius:12px;padding:12px}"
        ".lsec{font-size:.63rem;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.7px;text-align:center;margin-bottom:7px}"
        ".donut-wrap{position:relative;width:120px;height:120px;margin:0 auto 9px}"
        ".donut-center{position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);text-align:center;pointer-events:none}"
        ".dpct{font-size:1.45rem;font-weight:800;line-height:1}"
        ".dlbl{font-size:.58rem;font-weight:700;margin-top:2px}"
        ".leg{display:flex;flex-direction:column;gap:4px;margin-bottom:8px}"
        ".leg-r{display:flex;align-items:center;gap:6px;font-size:.7rem}"
        ".ldot{width:8px;height:8px;border-radius:2px;flex-shrink:0}"
        ".oldest-date{font-size:.6rem;color:#475569;text-align:center;margin-bottom:6px}"
        ".divider{border:none;border-top:1px solid #1e293b;margin:7px 0}"
        ".tipe-hdr{display:grid;grid-template-columns:1fr 30px 22px 22px;font-size:.57rem;color:#334155;font-weight:700;margin-bottom:2px;padding:0 2px}"
        ".tipe-row{display:grid;grid-template-columns:1fr 30px 22px 22px;align-items:center;"
        "padding:3px 2px;border-bottom:1px solid #1e293b18;font-size:.68rem}"
        ".tipe-row:last-child{border:none}"
        ".mdot{display:inline-block;width:6px;height:6px;border-radius:50%;margin-right:4px}"
        ".media-sec{font-size:.63rem;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.5px;margin:0 0 4px}"
        ".mhdr,.mrow{display:grid;grid-template-columns:1fr 24px 20px 20px;gap:3px;align-items:center;padding:0 5px}"
        ".mhdr{font-size:.55rem;color:#334155;font-weight:700;margin-bottom:2px}"
        ".mhdr span:not(:first-child){text-align:center}"
        ".mrow{background:#080d1a;border-radius:4px;padding:4px 5px;margin-bottom:2px}"
        ".msrc{font-size:.6rem;color:#cbd5e1;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}"
        ".mtot{font-size:.65rem;font-weight:700;color:#e2e8f0;text-align:center}"
        ".mneg{font-size:.6rem;color:#f87171;text-align:center}"
        ".mpos{font-size:.6rem;color:#4ade80;text-align:center}"
        ".src-scroll{max-height:320px;overflow-y:auto}"
        ".art-panel{background:#0f172a;border:1px solid #1e293b;border-radius:12px;padding:12px}"
        ".art-hdr{font-size:.83rem;font-weight:700;margin-bottom:9px;padding-bottom:7px;border-bottom:1px solid #1e293b;text-align:center}"
        ".acard{display:flex;gap:9px;padding:7px 0;border-bottom:1px solid #ffffff06;align-items:flex-start}"
        ".acard:last-child{border:none}"
        ".ascore{font-size:1.3rem;font-weight:800;min-width:28px;text-align:center;line-height:1}"
        ".acontent{flex:1;min-width:0}"
        ".adate{font-size:.57rem;color:#334155;margin-bottom:2px}"
        ".atitle-link{display:block;font-size:.77rem;font-weight:600;color:#cbd5e1;text-decoration:none;"
        "line-height:1.4;margin-bottom:2px;transition:color .15s}"
        ".atitle-link:hover{color:#93c5fd;text-decoration:underline;text-underline-offset:2px}"
        ".atitle-link::after{content:'\\2197';font-size:.65em;opacity:0;margin-left:3px;color:#60a5fa;transition:opacity .15s}"
        ".atitle-link:hover::after{opacity:1}"
        ".atitle-nolink{display:block;font-size:.77rem;font-weight:600;color:#374151;line-height:1.4;margin-bottom:2px}"
        ".asrc{font-size:.63rem;color:#475569;margin-bottom:2px}"
        ".areason{font-size:.6rem;color:#64748b;font-style:italic}"
        ".abar{background:#1e293b;border-radius:3px;height:3px;overflow:hidden;margin-top:3px}"
        ".afill{height:100%;border-radius:3px}"
        ".alink{display:inline-block;margin-top:4px;font-size:.61rem;color:#60a5fa;text-decoration:none;"
        "padding:2px 7px;border:1px solid #1d4ed8;border-radius:10px}"
        ".alink:hover{background:#1d4ed822;color:#93c5fd}"
        ".empty{color:#334155;text-align:center;padding:20px;font-style:italic;font-size:.78rem}"
        ".narrative-box{background:#0f172a;border:1px solid #1e293b;border-left:4px solid #3b82f6;"
        "border-radius:12px;padding:14px 18px;margin-bottom:11px}"
        ".ntitle{font-size:.8rem;font-weight:700;color:#93c5fd;margin-bottom:6px}"
        ".narrative-text{font-size:.77rem;line-height:1.8;color:#94a3b8}"
        ".footer{text-align:center;color:#1e293b;font-size:.65rem;padding-top:10px;border-top:1px solid #1e293b}"
        "@media(min-width:1400px){.main-grid{grid-template-columns:230px 1fr 1fr}}"
    )

    # ── JavaScript Logic (regular string — no f-string escaping needed) ───────
    js_logic = """
const TC={"News":"#58A6FF","Blog":"#4ade80","Twitter":"#1D9BF0","Podcast":"#BC8CFF","Google Alerts":"#F0A500","E-Commerce":"#f87171"};
let ALL=EMBEDDED;

async function init(){
  try{const r=await fetch('./articles.json?v='+Date.now());if(r.ok)ALL=await r.json();}catch(e){}
  render(0);
}

function setFilter(btn,days){
  document.querySelectorAll('.filter-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  render(days);
}

function getCutoff(days){
  const d=new Date();d.setDate(d.getDate()-days);return d.toISOString().split('T')[0];
}

function render(days){
  const today=new Date().toISOString().split('T')[0];
  // days=0 → hanya artikel hari ini
  const arts=days===0?ALL.filter(a=>a.run_date===today):ALL.filter(a=>a.run_date>=getCutoff(days));
  // sort: score DESC, lalu tanggal ASC (kasus lama dgn skor sama muncul lebih dulu)
  const neg=arts.filter(a=>a.sentiment==='negatif').sort((a,b)=>b.score-a.score||(a.run_date<b.run_date?-1:1));
  const pos=arts.filter(a=>a.sentiment==='positif').sort((a,b)=>b.score-a.score||(b.run_date<a.run_date?-1:1));
  const tot=arts.length,nN=neg.length,nP=pos.length;
  const pP=tot?Math.round(nP/tot*100):0,pN=tot?Math.round(nN/tot*100):0;
  const sentColor=pN>pP?'#f87171':(pP>pN?'#4ade80':'#fbbf24');
  const sentLabel=pN>pP?'NEGATIF ⚠️':(pP>pN?'POSITIF ✅':'NETRAL ⚡');

  // filter info
  const dates=[...new Set(arts.map(a=>a.run_date))].sort();
  const lbls={0:'Hari ini',1:'Kemarin',7:'7 hari',30:'30 hari',90:'3 bulan',180:'6 bulan',365:'1 tahun'};
  const info=days===0?(dates.length?'Hari ini · '+arts.length+' artikel':'Belum ada data hari ini')
    :days===1?(dates.length?'Data: '+fd(dates[dates.length-1]):'Tidak ada data')
    :(lbls[days]||days+' hari')+' terakhir · '+dates.length+' hari data';
  gi('filter-info').textContent=info;

  // stats cards
  const srcSet=new Set(arts.map(a=>a.source).filter(Boolean));
  gi('s-total').textContent=tot.toLocaleString('id');
  gi('s-neg').textContent=nN.toLocaleString('id');
  gi('s-pos').textContent=nP.toLocaleString('id');
  gi('s-pct').textContent=pP+'%';gi('s-pct').style.color=sentColor;
  gi('s-sources').textContent=srcSet.size.toLocaleString('id');

  // donut & legend
  gi('dpct').textContent=pP+'%';gi('dpct').style.color=sentColor;
  gi('dlbl').textContent=sentLabel;gi('dlbl').style.color=sentColor;
  gi('leg-neg').textContent='Negatif '+nN+' ('+pN+'%)';
  gi('leg-pos').textContent='Positif '+nP+' ('+pP+'%)';
  drawDonut(pP);
  if(dates.length){
    const d0=dates[0],d1=dates[dates.length-1];
    gi('oldest-date').textContent=d0===d1?'📅 '+fd(d0):'📅 '+fd(d0)+' – '+fd(d1);
  }else gi('oldest-date').textContent='';

  // per tipe media
  const td={};
  arts.forEach(a=>{const m=a.media_type||'News';if(!td[m])td[m]={t:0,n:0,p:0};td[m].t++;
    if(a.sentiment==='negatif')td[m].n++;if(a.sentiment==='positif')td[m].p++;});
  gi('tipe-rows').innerHTML=Object.entries(td).sort((a,b)=>b[1].t-a[1].t).map(([m,d])=>{
    const c=TC[m]||'#8B949E';
    return `<div class="tipe-row"><span><span class="mdot" style="background:${c}"></span>`
      +`<span style="color:${c}">${esc(m)}</span></span>`
      +`<span style="text-align:center;color:#e2e8f0">${d.t}</span>`
      +`<span style="text-align:center;color:#f87171">${d.n||'·'}</span>`
      +`<span style="text-align:center;color:#4ade80">${d.p||'·'}</span></div>`;
  }).join('');

  // media sources
  const sc={},sn={},sp={};
  arts.forEach(a=>{const s=a.source;if(!s)return;sc[s]=(sc[s]||0)+1;
    if(a.sentiment==='negatif')sn[s]=(sn[s]||0)+1;
    if(a.sentiment==='positif')sp[s]=(sp[s]||0)+1;});
  gi('src-scroll').innerHTML=Object.entries(sc).sort((a,b)=>b[1]-a[1]).slice(0,20).map(([s,c])=>
    `<div class="mrow"><span class="msrc" title="${esc(s)}">${esc(s)}</span>`
    +`<span class="mtot">${c}</span><span class="mneg">${sn[s]||'·'}</span>`
    +`<span class="mpos">${sp[s]||'·'}</span></div>`
  ).join('');

  // articles — max 2 artikel per tanggal agar tersebar lintas periode
  function topArts(sorted, n, maxPerDate){
    const out=[],cnt={};
    for(const a of sorted){
      const d=a.run_date||'';
      if((cnt[d]||0)<maxPerDate){out.push(a);cnt[d]=(cnt[d]||0)+1;}
      if(out.length>=n)break;
    }
    return out;
  }
  const mpd=days<=7?10:2; // periode pendek: bebas; periode panjang: max 2/hari
  renderArts(topArts(neg,15,mpd),'neg-list','neg-count','#f87171');
  renderArts(topArts(pos,15,mpd),'pos-list','pos-count','#4ade80');

  // narrative — dinamis sesuai periode
  const oldest=dates[0]||'',newest=dates[dates.length-1]||'';
  if((days===0||days===1)&&NAR){
    gi('nar-day').textContent=fd(RUN_DATE);
    gi('nar-text').innerHTML=NAR;
  }else if(arts.length){
    const lblPeriode={7:'7 hari',30:'30 hari',90:'3 bulan',180:'6 bulan',365:'1 tahun'}[days]||(days+' hari');
    const dateRange=oldest===newest?fd(oldest):(fd(oldest)+' – '+fd(newest));
    const srcC={};arts.forEach(a=>{if(a.source)srcC[a.source]=(srcC[a.source]||0)+1;});
    const topSrc=Object.entries(srcC).sort((a,b)=>b[1]-a[1]).slice(0,3).map(([s,c])=>s+' ('+c+')').join(', ');
    let nar='Dalam <strong>'+lblPeriode+'</strong> terakhir ('+dateRange+'), terpantau '
      +'<strong>'+tot+'</strong> artikel berita tentang Bank Mandiri dari '
      +'<strong>'+Object.keys(srcC).length+'</strong> sumber media.<br><br>';
    nar+='Sentimen: <span style="color:#4ade80;font-weight:600">'+pP+'% positif</span> ('+nP+' artikel) '
      +'dan <span style="color:#f87171;font-weight:600">'+pN+'% negatif</span> ('+nN+' artikel).<br><br>';
    if(topSrc)nar+='Media paling aktif: '+esc(topSrc)+'.<br><br>';
    if(neg.length)nar+='Berita negatif skor tertinggi: <em>“'+esc(neg[0].title)+'”</em>'
      +' — '+esc(neg[0].source)+' ('+fd(neg[0].run_date)+').<br>';
    if(pos.length)nar+='Berita positif skor tertinggi: <em>“'+esc(pos[0].title)+'”</em>'
      +' — '+esc(pos[0].source)+' ('+fd(pos[0].run_date)+').';
    gi('nar-day').textContent=dateRange;
    gi('nar-text').innerHTML=nar;
  }
}

function renderArts(items,listId,countId,color){
  gi(countId).textContent='('+items.length+' artikel)';
  if(!items.length){gi(listId).innerHTML='<div class="empty">Tidak ada berita dalam kategori ini</div>';return;}
  gi(listId).innerHTML=items.map(a=>{
    const w=(a.score/10*100).toFixed(0);
    const d=a.run_date?fd(a.run_date):'';
    const th=a.link?`<a class="atitle-link" href="${esc(a.link)}" target="_blank" rel="noopener">${esc(a.title)}</a>`
      :`<span class="atitle-nolink">${esc(a.title)}</span>`;
    const lh=a.link?`<a class="alink" href="${esc(a.link)}" target="_blank" rel="noopener">↗ Buka artikel</a>`:'';
    return `<div class="acard"><div class="ascore" style="color:${color}">${a.score}</div>`
      +`<div class="acontent"><div class="adate">${d}</div>${th}`
      +`<div class="asrc">${esc(a.source||'')} <span class="areason">${esc(a.reason||'')}</span></div>`
      +`<div class="abar"><div class="afill" style="width:${w}%;background:${color}"></div></div>${lh}</div></div>`;
  }).join('');
}

function drawDonut(pctPos){
  const cv=document.getElementById('donutCanvas');if(!cv)return;
  const ctx=cv.getContext('2d');
  const cx=60,cy=60,r=54,ir=34,gap=0.05,s0=-Math.PI/2;
  const pn=(100-pctPos)/100,pp=pctPos/100;
  function arc(s,en,col){ctx.beginPath();ctx.moveTo(cx,cy);ctx.arc(cx,cy,r,s,en);
    ctx.arc(cx,cy,ir,en,s,true);ctx.closePath();ctx.fillStyle=col;ctx.fill();}
  ctx.fillStyle='#0f172a';ctx.fillRect(0,0,120,120);
  if(pn>0.01)arc(s0,s0+Math.PI*2*pn-gap,'#f87171');
  if(pp>0.01)arc(s0+Math.PI*2*pn+gap,s0+Math.PI*2-gap,'#4ade80');
}

function gi(id){return document.getElementById(id);}
function esc(s){return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');}
function fd(iso){if(!iso)return'';const[y,m,d]=iso.split('-');
  const mn=['','Jan','Feb','Mar','Apr','Mei','Jun','Jul','Agu','Sep','Okt','Nov','Des'];
  return parseInt(d)+' '+(mn[parseInt(m)]||m)+' '+y;}

init();
"""

    # ── Build HTML via string concatenation (avoids f-string brace conflicts) ─
    html = (
        '<!DOCTYPE html>\n<html lang="id">\n<head>\n'
        '<meta charset="UTF-8">\n'
        '<meta name="viewport" content="width=device-width,initial-scale=1">\n'
        '<meta http-equiv="refresh" content="600">\n'
        + f'<title>Media Monitoring BMRI — {date_str}</title>\n'
        + '<style>\n' + css + '\n</style>\n</head>\n<body>\n\n'
        + '<header class="hdr">\n'
        '  <h1>\U0001f4ca Media Monitoring Bank Mandiri (BMRI)</h1>\n'
        + f'  <div class="sub">Sumber: Talkwalker \xb7 Google Alerts \xb7 Podcast'
          f' &nbsp;\xb7&nbsp; Auto-refresh setiap 10 menit</div>\n'
        + f'  <div class="upd">Terakhir diperbarui: {now_str}</div>\n'
        + '</header>\n\n'
        '<div class="click-legend">\n'
        '  <span><span class="cl-dot" style="background:#cbd5e1"></span>'
        'Judul terang = ada link, bisa diklik</span>\n'
        '  <span><span class="cl-dot" style="background:#374151"></span>'
        'Judul abu-abu = tidak ada link</span>\n'
        '</div>\n\n'
        '<div class="filter-bar">\n'
        '  <span class="filter-label">⏱ Periode:</span>\n'
        '  <button class="filter-btn active" onclick="setFilter(this,0)">Hari Ini</button>\n'
        '  <button class="filter-btn" onclick="setFilter(this,1)">Kemarin</button>\n'
        '  <button class="filter-btn" onclick="setFilter(this,7)">7 Hari</button>\n'
        '  <button class="filter-btn" onclick="setFilter(this,30)">30 Hari</button>\n'
        '  <button class="filter-btn" onclick="setFilter(this,90)">3 Bulan</button>\n'
        '  <button class="filter-btn" onclick="setFilter(this,180)">6 Bulan</button>\n'
        '  <button class="filter-btn" onclick="setFilter(this,365)">1 Tahun</button>\n'
        '  <div class="filter-divider"></div>\n'
        '  <span class="filter-info" id="filter-info">Memuat data...</span>\n'
        '</div>\n\n'
        '<div class="stats-row">\n'
        '  <div class="stat-card"><div class="stat-num" style="color:#94a3b8" id="s-total">—</div>'
        '<div class="stat-lbl">Total Artikel</div></div>\n'
        '  <div class="stat-card"><div class="stat-num" style="color:#f87171" id="s-neg">—</div>'
        '<div class="stat-lbl">Negatif</div></div>\n'
        '  <div class="stat-card"><div class="stat-num" style="color:#4ade80" id="s-pos">—</div>'
        '<div class="stat-lbl">Positif</div></div>\n'
        '  <div class="stat-card"><div class="stat-num" id="s-pct">—</div>'
        '<div class="stat-lbl">Sentimen Positif</div></div>\n'
        '  <div class="stat-card"><div class="stat-num" style="color:#60a5fa" id="s-sources">—</div>'
        '<div class="stat-lbl">Media Sumber</div></div>\n'
        '</div>\n\n'
        '<div class="main-grid">\n\n'
        '  <div class="left-panel">\n'
        '    <div class="lsec">Sentimen Keseluruhan</div>\n'
        '    <div class="donut-wrap">\n'
        '      <canvas id="donutCanvas" width="120" height="120"></canvas>\n'
        '      <div class="donut-center">'
        '<div class="dpct" id="dpct">—</div>'
        '<div class="dlbl" id="dlbl">Memuat...</div></div>\n'
        '    </div>\n'
        '    <div class="leg">\n'
        '      <div class="leg-r"><div class="ldot" style="background:#f87171"></div>'
        '<span style="color:#f87171" id="leg-neg">Negatif —</span></div>\n'
        '      <div class="leg-r"><div class="ldot" style="background:#4ade80"></div>'
        '<span style="color:#4ade80" id="leg-pos">Positif —</span></div>\n'
        '    </div>\n'
        '    <div class="oldest-date" id="oldest-date"></div>\n'
        '    <hr class="divider">\n'
        '    <div class="tipe-hdr"><span>Tipe</span>'
        '<span style="text-align:center">Ttl</span>'
        '<span style="color:#f87171;text-align:center">N</span>'
        '<span style="color:#4ade80;text-align:center">P</span></div>\n'
        '    <div id="tipe-rows"></div>\n'
        '    <hr class="divider">\n'
        '    <div class="media-sec">Media Terpantau</div>\n'
        '    <div class="mhdr"><span>Sumber</span>'
        '<span style="text-align:center">Ttl</span>'
        '<span style="color:#f87171;text-align:center">N</span>'
        '<span style="color:#4ade80;text-align:center">P</span></div>\n'
        '    <div class="src-scroll" id="src-scroll"></div>\n'
        '  </div>\n\n'
        '  <div class="art-panel">\n'
        '    <div class="art-hdr" style="color:#f87171">▼ NEGATIF'
        ' &nbsp;<span id="neg-count"></span></div>\n'
        '    <div id="neg-list"></div>\n'
        '  </div>\n\n'
        '  <div class="art-panel">\n'
        '    <div class="art-hdr" style="color:#4ade80">▲ POSITIF'
        ' &nbsp;<span id="pos-count"></span></div>\n'
        '    <div id="pos-list"></div>\n'
        '  </div>\n\n'
        '</div>\n\n'
        '<div class="narrative-box">\n'
        '  <div class="ntitle">\U0001f4cb Ringkasan Naratif Eksekutif'
        ' &nbsp;<small id="nar-day" style="font-weight:400;color:#64748b;font-size:.68rem"></small></div>\n'
        '  <div class="narrative-text" id="nar-text"></div>\n'
        '</div>\n\n'
        '<footer class="footer">\n'
        '  Powered by Claude AI (Anthropic) &nbsp;\xb7&nbsp;'
        ' Talkwalker \xb7 Google Alerts \xb7 ListenNotes\n'
        '  &nbsp;\xb7&nbsp; Data historis tersimpan hingga 1 tahun'
        ' &nbsp;\xb7&nbsp; Update otomatis pukul 07.00 WIB\n'
        '</footer>\n\n'
        '<script>\n'
        + f'const EMBEDDED={embedded_json};\n'
        + f'const NAR=`{nar_safe}`;\n'
        + f"const RUN_DATE='{run_date}';\n"
        + js_logic
        + '</script>\n</body>\n</html>'
    )
    return html


def send_email(articles, date_str, chart_path, narrative=None):
    """Kirim email HTML dengan grafik embedded + lampiran Excel ke EMAIL_RECIPIENT."""
    positif = len([a for a in articles if a.get("sentiment") == "positif"])
    negatif = len([a for a in articles if a.get("sentiment") == "negatif"])

    msg = MIMEMultipart("mixed")
    msg["Subject"] = f"📊 Media Monitoring BMRI — {date_str} | {positif} Positif / {negatif} Negatif"
    msg["From"]    = EMAIL_SENDER
    msg["To"]      = EMAIL_RECIPIENT

    related = MIMEMultipart("related")
    alternative = MIMEMultipart("alternative")
    html_body = build_html_email(articles, date_str, chart_path, narrative=narrative)
    alternative.attach(MIMEText(html_body, "html", "utf-8"))
    related.attach(alternative)

    with open(chart_path, "rb") as f:
        img = MIMEImage(f.read(), _subtype="png")
    img.add_header("Content-ID", "<sentiment_chart>")
    img.add_header("Content-Disposition", "inline", filename="sentiment_chart.png")
    related.attach(img)
    msg.attach(related)

    excel_path = None
    try:
        excel_path = generate_excel_report(articles, date_str)
        filename   = f"MediaMonitoring_BMRI_{date_str.replace(' ','_')}.xlsx"
        with open(excel_path, "rb") as f:
            excel_data = f.read()
        excel_part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        excel_part.set_payload(excel_data)
        encoders.encode_base64(excel_part)
        excel_part.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(excel_part)
        print(f"[Email] Excel terlampir: {filename}")
    except Exception as e:
        print(f"[Email] Gagal buat Excel: {e}")
    finally:
        if excel_path:
            try:
                os.remove(excel_path)
            except Exception:
                pass

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=30) as smtp:
            smtp.login(EMAIL_SENDER, EMAIL_APP_PASS)
            smtp.send_message(msg)
        print(f"[Email] Terkirim ke {EMAIL_RECIPIENT}")
    except Exception as e:
        print(f"[Email] Error: {e}")
        raise


# ─────────────────────────────────────────────
# 9. MAIN
# ─────────────────────────────────────────────

# ─────────────────────────────────────────────
# 9. ONLINE NEWS FETCHER (integrated)
# ─────────────────────────────────────────────
def fetch_online_news_articles(existing_seen):
    import urllib.parse as _up, urllib.request as _ur
    import xml.etree.ElementTree as _ET, time as _t, re as _re

    HDR  = {"User-Agent": "Mozilla/5.0 Chrome/120.0.0.0"}
    CUT  = (datetime.date.today() - datetime.timedelta(days=2)).strftime("%Y-%m-%d")
    KWLO = ["bank mandiri", "bmri"]

    PORTALS = [
        ("Kompas.com",    "https://rss.kompas.com/money/read/xml/indeks/1/"),
        ("Detik Finance", "https://finance.detik.com/rss"),
        ("Bisnis.com",    "https://feeds.bisnis.com/bisnis/rss/finansial/keuangan"),
        ("Kontan",        "https://www.kontan.co.id/rss/news.rss"),
        ("CNBC Indonesia","https://www.cnbcindonesia.com/rss"),
        ("Tribun Bisnis", "https://www.tribunnews.com/rss/bisnis"),
        ("IDX Channel",   "https://www.idxchannel.com/feed"),
        ("Investor Daily","https://investor.id/feed"),
        ("Tempo Money",   "https://rss.tempo.co/bisnis"),
        ("Republika",     "https://rss.republika.co.id/rss/ekonomi/keuangan"),
    ]
    GNEWS_KWS = [
        "Bank Mandiri", "BMRI", "saham BMRI", "laba Bank Mandiri",
        "dividen BMRI", "Livin Mandiri", "kredit Mandiri",
        "Bank Mandiri OJK", "Bank Mandiri kasus",
        "Bank Mandiri tersangka", "Bank Mandiri dilaporkan",
        "Darmawan Junaidi Bank Mandiri", "RUPS Bank Mandiri",
    ]
    BLOG_KWS = [
        "Bank Mandiri Kaskus", "BMRI forum saham",
        "Bank Mandiri pengalaman nasabah", "Livin Mandiri masalah",
    ]

    def _dom(u):
        try: return _up.urlparse(u).netloc.replace("www.","")
        except: return ""

    def _date(d):
        if not d: return None
        m = _re.search(r"(\d{4}-\d{2}-\d{2})", str(d))
        return m.group(1) if m else None

    def _art(title, src, url, pub, mtype, snip=""):
        return {"title":title,"source":src,"url":url,"run_date":pub,
                "media_type":mtype,"snippet":snip[:200],"summary":"",
                "sentiment":"netral","sentiment_label":"netral",
                "score":3,"topic":"Lainnya","reason":""}

    def _rss(url, src_name, mtype="News", kw_filter=True):
        out = []
        try:
            req = _ur.Request(url, headers=HDR)
            with _ur.urlopen(req, timeout=12) as r:
                root = _ET.fromstring(r.read())
            for item in root.findall(".//item"):
                title = (item.findtext("title") or "").strip()
                link  = item.findtext("link") or ""
                desc  = (item.findtext("description") or "")[:300]
                pub   = _date(item.findtext("pubDate") or "")
                se    = item.find("source")
                src   = (se.text.strip() if se is not None and se.text else "") or src_name or _dom(link)
                if not title or not pub or pub < CUT: continue
                if kw_filter and not any(k in (title+" "+desc).lower() for k in KWLO): continue
                key = title[:80].lower()
                if key not in existing_seen:
                    existing_seen.add(key)
                    out.append(_art(title, src, link, pub, mtype, desc))
        except Exception as e:
            print(f"    RSS {src_name}: {e}")
        return out

    results = []
    print("  [Online] Google News RSS...")
    for kw in GNEWS_KWS:
        q = _up.quote(kw)
        results.extend(_rss(f"https://news.google.com/rss/search?q={q}&hl=id&gl=ID&ceid=ID:id",
                            "Google News","News",False))
        _t.sleep(0.5)

    print("  [Online] Portal berita...")
    for name, url in PORTALS:
        results.extend(_rss(url, name, "News", True))
        _t.sleep(0.3)

    print("  [Online] Blog & Forum...")
    for kw in BLOG_KWS:
        q = _up.quote(kw)
        results.extend(_rss(f"https://news.google.com/rss/search?q={q}&hl=id&gl=ID&ceid=ID:id",
                            "Blog/Forum","Blog",False))
        _t.sleep(0.5)

    yt_key = os.environ.get("YOUTUBE_API_KEY","")
    if yt_key:
        print("  [Online] YouTube...")
        for kw in ["Bank Mandiri","BMRI saham","Livin Mandiri"]:
            try:
                url = (f"https://www.googleapis.com/youtube/v3/search?part=snippet"
                       f"&q={_up.quote(kw)}&type=video&order=date&maxResults=20"
                       f"&key={yt_key}&publishedAfter={CUT}T00:00:00Z")
                with _ur.urlopen(_ur.Request(url,headers=HDR),timeout=12) as r:
                    data = json.loads(r.read())
                for item in data.get("items",[]):
                    sn = item.get("snippet",{})
                    t2 = sn.get("title","").strip()
                    vid= item.get("id",{}).get("videoId","")
                    pub= sn.get("publishedAt","")[:10] or datetime.date.today().strftime("%Y-%m-%d")
                    key= t2[:80].lower()
                    if t2 and vid and key not in existing_seen:
                        existing_seen.add(key)
                        results.append(_art(t2,sn.get("channelTitle","YouTube"),
                            f"https://youtube.com/watch?v={vid}",pub,"YouTube",
                            sn.get("description","")[:200]))
                _t.sleep(1)
            except Exception as e:
                print(f"    YouTube '{kw}': {e}")

    xb = os.environ.get("X_BEARER_TOKEN","")
    if xb:
        print("  [Online] Twitter/X...")
        for kw in ["Bank Mandiri","BMRI"]:
            try:
                url = (f"https://api.twitter.com/2/tweets/search/recent"
                       f"?query={_up.quote(kw+' lang:id -is:retweet')}"
                       f"&max_results=50&tweet.fields=created_at,author_id"
                       f"&expansions=author_id&user.fields=name")
                with _ur.urlopen(_ur.Request(url,headers={"Authorization":f"Bearer {xb}",**HDR}),timeout=12) as r:
                    data = json.loads(r.read())
                users = {u["id"]:u.get("name","Twitter") for u in data.get("includes",{}).get("users",[])}
                for tw in data.get("data",[]):
                    title = tw.get("text","").replace("\n"," ").strip()
                    pub   = tw.get("created_at","")[:10] or datetime.date.today().strftime("%Y-%m-%d")
                    key   = title[:80].lower()
                    if key not in existing_seen:
                        existing_seen.add(key)
                        results.append(_art(title,users.get(tw.get("author_id",""),"Twitter"),
                            f"https://twitter.com/i/web/status/{tw['id']}",pub,"Twitter",title))
                _t.sleep(1)
            except Exception as e:
                print(f"    Twitter '{kw}': {e}")

    print(f"  [Online] +{len(results)} artikel baru")
    return results


# ─────────────────────────────────────────────
# 10. MAIN
# ─────────────────────────────────────────────
def main():
    now      = datetime.datetime.utcnow() + datetime.timedelta(hours=7)  # UTC → WIB
    date_str = now.strftime("%d %B %Y %H:%M WIB")
    print(f"\n{'='*60}")
    print(f"  BANK MANDIRI MEDIA MONITORING — {date_str}")
    print(f"{'='*60}\n")

    # 1. Gmail
    print("[1/5] Gmail (Talkwalker + Google Alerts)...")
    try:
        service   = get_gmail_service()
        tw_bodies = fetch_talkwalker_emails(service)
        ga_bodies = fetch_google_alerts_emails(service)
    except Exception as e:
        print(f"  Gmail error: {e}")
        tw_bodies, ga_bodies = [], []

    articles_tw = parse_articles(tw_bodies) if tw_bodies else []
    articles_ga = parse_google_alerts(ga_bodies) if ga_bodies else []

    try:
        articles_pod = fetch_podcast_mentions()
    except Exception as e:
        print(f"  Podcast error: {e}")
        articles_pod = []

    # 2. Dedup awal
    seen_titles = set()
    articles    = []
    for a in articles_tw + articles_ga + articles_pod:
        key = a.get("title","")[:80].lower()
        if key not in seen_titles:
            seen_titles.add(key)
            articles.append(a)

    # 3. Online news
    print("[2/5] Online news, RSS, portal, YouTube, Twitter...")
    try:
        articles.extend(fetch_online_news_articles(seen_titles))
    except Exception as e:
        print(f"  Online news error: {e}")

    n_on = len(articles)-len(articles_tw)-len(articles_ga)-len(articles_pod)
    print(f"Total: {len(articles)} artikel "
          f"(TW:{len(articles_tw)} GA:{len(articles_ga)} Pod:{len(articles_pod)} Online:{n_on})")

    if not articles:
        print("Tidak ada artikel.")
        return

    # 4. AI analysis
    print(f"[3/5] Analisis AI ({len(articles)} artikel)...")
    articles = analyze_sentiment(articles)

    # 5. Narasi & grafik
    print("[4/5] Narasi & grafik...")
    narrative = generate_narrative_summary(articles, date_str)
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        chart_path = tmp.name
    generate_chart(articles, chart_path)

    # 6. Dashboard + Email (Telegram: DIMATIKAN)
    print("[5/5] Dashboard & Email...")
    try:
        docs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "docs")
        os.makedirs(docs_dir, exist_ok=True)
        save_articles_history(articles, date_str, narrative, docs_dir)
        html = generate_tv_dashboard(articles, date_str, chart_path, narrative=narrative)
        with open(os.path.join(docs_dir, "index.html"), "w", encoding="utf-8") as f:
            f.write(html)
        print("  Dashboard diperbarui")
    except Exception as e:
        print(f"  Dashboard error: {e}")

    try:
        send_email(articles, date_str, chart_path, narrative=narrative)
    except Exception as e:
        print(f"  Email error: {e}")

    try: os.unlink(chart_path)
    except: pass

    print(f"\nSelesai! {len(articles)} artikel dianalisis.")


if __name__ == "__main__":
    main()
